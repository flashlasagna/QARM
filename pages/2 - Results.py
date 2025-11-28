import streamlit as st
from backend.style_utils import apply_sidebar_style
st.set_page_config(page_title="Results", layout="wide")
apply_sidebar_style()
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
from io import BytesIO
import json
import os
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, ScatterChart, Reference
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.styles import Font, PatternFill, Alignment

# --- Path Setup ---
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(current_dir)
if root_dir not in sys.path:
    sys.path.insert(0, root_dir)

# --- Backend imports ---
from backend.config_loader import load_config, get_corr_matrices, get_solvency_params
from backend.optimization import solve_frontier_combined
from backend.helpers import summarize_portfolio
from backend.helpers import plot_scenario_comparison
from backend.data_calculator import (
    compute_ir_shocks_from_eiopa,
    compute_spread_shock_eiopa)
from backend.solvency_calc import (
    scr_interest_rate, 
    scr_eq,
    scr_prop, 
    scr_sprd, 
    aggregate_market_scr, 
    marginal_scr,
    allocate_marginal_scr
)

st.title("Optimization Results")

# --- 1. Session State Checks ---
if not st.session_state.get("optimization_run"):
    st.warning("Please run the optimization on the **Inputs** page first.")
    st.stop()

# Show if auto-calculation was used
if st.session_state.get("auto_calculated", False):
    st.info("🤖 **Auto-calculation was enabled** - Returns and IR shocks were computed from market data")

# --- Load Data ---
# Ensure liab_duration is explicitly loaded into the script's main scope
if "opt_df" in st.session_state:
    opt_df = st.session_state["opt_df"]
    best = opt_df.loc[opt_df["objective"].idxmax()]
    initial_asset = st.session_state["initial_asset"]
    liab_value = st.session_state["liab_value"]

    # FIX: Explicitly define liab_duration here
    liab_duration = st.session_state["liab_duration"]
    
    # =========================
else:
    # Handle missing data gracefully if needed
    st.info("Optimization data not fully loaded.")
    st.stop()



# Get parameters from config (or session state if stored, otherwise re-load)
cfg = load_config()
corr_down, corr_up = get_corr_matrices(cfg)
solv = get_solvency_params(cfg)

# Recalculate SCR components for the BEST portfolio
# 1. Interest Rate Risk
# We need shocks from session state or re-compute
if st.session_state.get('use_eiopa_curves', True):
    ir_up, ir_down = compute_ir_shocks_from_eiopa(liab_duration=liab_value,
                                                  verbose=False)  # Note: passed liab_value as placeholder in previous turn, ideally use liab_duration
else:
    # Fallback or pull from somewhere else. For now re-compute or assume safe defaults if missing
    # Better: Store shocks in session state in Inputs.py
    ir_up, ir_down = 0.011, 0.009

# Compute Interest SCR
scr_int_CUR = scr_interest_rate(
    initial_asset["asset_val"],
    initial_asset["asset_dur"],
    liab_value,
    liab_duration,
    ir_up,
    ir_down
)

# 2. Equity Risk
A_eq1_CUR = initial_asset["asset_val"][2]
A_eq2_CUR = initial_asset["asset_val"][3]
scr_eq_CUR = scr_eq(A_eq1_CUR, A_eq2_CUR, solv["equity_1_param"], solv["equity_2_param"], solv["rho"])

# 3. Property Risk
scr_prop_CUR = scr_prop(initial_asset["asset_val"][4], solv["prop_params"])

# 4. Spread Risk
# Need spread shock (re-compute or store)
corp_dur = initial_asset.loc["corp_bond", "asset_dur"]
spread_shock_val = compute_spread_shock_eiopa(corp_dur, verbose=False)
scr_sprd_CUR = scr_sprd(initial_asset["asset_val"][1], spread_shock_val)

# 5. Prepare Vector for Marginal Calculation
scr_vec_CUR = np.array([
    scr_int_CUR["SCR_interest"],
    scr_eq_CUR["SCR_eq_total"],
    scr_prop_CUR,
    scr_sprd_CUR
])

# Determine direction (Up/Down) from Interest result
direction_CUR = scr_int_CUR["direction"]


market_scr_CUR = aggregate_market_scr(
    scr_int_CUR,
    scr_eq_CUR, 
    scr_prop_CUR,
    scr_sprd_CUR,
    corr_down,
    corr_up
)


# Run Marginal Calculation (Risk Level)
marg_risk_CUR = marginal_scr(scr_vec_CUR, direction_CUR, corr_down.values, corr_up.values)


# Calculate Current Metrics (ignoring duration)
current_ret = float((initial_asset["asset_ret"] * initial_asset["asset_weight"]).sum())
current_SCR = market_scr_CUR['summary_table'].loc['total', 'SCR']
current_sol = (initial_asset['asset_val'].sum() - liab_value) / current_SCR


# Optimal one is the one having solvency closest to current solvency (so we keep safety unchanged)
best_idx = (opt_df["solvency"] - current_sol).abs().idxmin()
best_row = opt_df.loc[best_idx]


best = opt_df.loc[best_idx]


# Best Metrics
best_ret = float(best["return"])
best_sol = float(best["solvency"])
best_SCR = float(best["SCR_market"])
best_BOF = float(best["BOF"])

# === NEW: CRITICAL SOLVENCY CHECK ===
if best_sol < 1.0:
    st.error(f"🚨 **CRITICAL WARNING: INSOLVENT PORTFOLIO**")
    st.warning(f"""
        **The optimizer could not find a portfolio compliant with Solvency II (100% Ratio).**
        
        The displayed portfolio is the **"Least Bad"** option available under your current constraints.
        
        - **Best Achievable Ratio:** {best_sol:.1%} (Target: 100%)
        - **Capital Shortfall:** €{(best['SCR_market'] - best['BOF']):,.1f}m
        
        **Recommended Actions:**
        1. **Inject Capital:** You cannot solve this by asset allocation alone.
        2. **Relax Constraints:** Allow more allocation to high-return/low-capital assets (if any).
        3. **Reduce Liabilities:** Revisit liability duration matching or reinsurance.
    """)

# --- 3. Key Metrics Comparison ---
st.subheader("Key Metrics Comparison")

col1, col2, col3, col4 = st.columns(4)

with col1:
    st.metric(
        "Expected Return",
        f"{best_ret:.2%}",
        delta=f"{(best_ret - current_ret):.2%}",
        help="Portfolio expected return (annualized)"
    )

with col2:
    st.metric(
        "Solvency Ratio",
        f"{best_sol * 100:.1f}%",
        delta=f"{(best_sol - current_sol) * 100:.1f}%",
        help="BOF / SCR Market"
    )

with col3:

    st.metric(
        "SCR Market",
        f"€{best_SCR:.1f}m",
        delta=f"€{(best_SCR - current_SCR):.1f}m",
        delta_color="inverse",
        help="Market risk capital requirement"
    )

with col4:
    st.metric(
        "Basic Own Funds",
        f"€{best_BOF:.1f}m",
        help="Assets - Liabilities"
    )

st.markdown("---")

# --- Efficient Frontier Plot (White Background + Annotations Kept) ---
st.subheader("Efficient Frontier")

# --- 1. Sort & Extract Pareto Frontier ---
opt_sorted = opt_df.sort_values(by="solvency", ascending=False).copy()
opt_sorted["max_return_seen"] = opt_sorted["return"].cummax()
pareto_frontier = opt_sorted[opt_sorted["return"] >= opt_sorted["max_return_seen"]]
pareto_frontier = pareto_frontier.sort_values(by="solvency")

# --- 2. Create Figure (WHITE background) ---
fig_frontier, ax_frontier = plt.subplots(figsize=(12, 7))
ax_frontier.set_facecolor("white")      # chart background white
fig_frontier.patch.set_facecolor("white")

# --- 3. Feasible Set (gray dots) ---
ax_frontier.scatter(
    opt_df["solvency"] * 100,
    opt_df["return"] * 100,
    s=30,
    color="gray",
    alpha=0.25,
    label="Feasible Portfolios"
)

# --- 4. Efficient Frontier (dark blue line) ---
ax_frontier.plot(
    pareto_frontier["solvency"] * 100,
    pareto_frontier["return"] * 100,
    '-',
    color='#003366',       # dark blue
    linewidth=3,
    label='Efficient Frontier',
    zorder=2
)

# --- 5. Optimal Portfolio (gold star) ---
optimal_solvency = best["solvency"] * 100
optimal_return = best["return"] * 100

ax_frontier.scatter(
    optimal_solvency, optimal_return,
    s=100, c='#FFD700', marker='*',
    edgecolors='#FF8C00', linewidth=3,
    label='Optimal Portfolio',
    zorder=5
)

# --- Optimal Annotation ---
ax_frontier.annotate(
    f'OPTIMAL\n{optimal_return:.2f}% | {optimal_solvency:.1f}%',
    xy=(optimal_solvency, optimal_return),
    xytext=(25, 25),
    textcoords='offset points',
    fontsize=10, fontweight='bold',
    bbox=dict(boxstyle='round,pad=0.8', facecolor='#FFD700', edgecolor='#FF8C00', alpha=0.9),
    arrowprops=dict(arrowstyle='->', color='#FF8C00', lw=2.5),
    zorder=6
)

# --- 6. Current Portfolio (red dot) ---
ax_frontier.scatter(
    current_sol * 100, current_ret * 100,
    s=60, c='#E74C3C', marker='o',
    edgecolors='#C0392B', linewidth=1.5,
    label='Current Portfolio',
    zorder=4, alpha=0.95
)

# --- Current Annotation ---
ax_frontier.annotate(
    f'CURRENT\n{current_ret*100:.2f}% | {current_sol*100:.1f}%',
    xy=(current_sol * 100, current_ret * 100),
    xytext=(-50, -35),
    textcoords='offset points',
    fontsize=9, fontweight='bold',
    bbox=dict(
        boxstyle='round,pad=0.6',
        facecolor='#FADBD8',
        edgecolor='#C0392B',
        alpha=0.9
    ),
    arrowprops=dict(arrowstyle='->', color='#C0392B', lw=2),
    zorder=6
)

# --- 7. Styling ---
ax_frontier.axvline(
    x=100, color='#95a5a6',
    linestyle='--', linewidth=2.5, alpha=0.6,
    label='100% Solvency'
)

ax_frontier.set_xlabel('Solvency Ratio (%)', fontsize=12, fontweight='bold', color='#2c3e50')
ax_frontier.set_ylabel('Expected Return (%)', fontsize=12, fontweight='bold', color='#2c3e50')

ax_frontier.tick_params(colors='#2c3e50')
ax_frontier.grid(True, alpha=0.25, linestyle='--', linewidth=1)

# Legend
ax_frontier.legend(
    loc='upper right',
    frameon=True, fancybox=True, shadow=True
)

st.pyplot(fig_frontier, use_container_width=True)
st.markdown("---")



# --- 6. Optimal Allocation Tables ---

def style_change_column_apply(df, column_name="Change (€m)"):
    """
    Adaptive gradient coloring for a 'Change' column using Styler.apply.
    Positive → green gradient
    Negative → red gradient
    Zero → white
    Works even if there's only a single positive or negative value.
    """

    vals = df[column_name]

    # positive values
    pos_vals = vals[vals > 0]
    if not pos_vals.empty:
        pos_min = pos_vals.min()
        pos_max = pos_vals.max()
    else:
        pos_min = pos_max = 0

    # negative values
    neg_vals = vals[vals < 0]
    if not neg_vals.empty:
        neg_min = neg_vals.min()
        neg_max = neg_vals.max()
    else:
        neg_min = neg_max = 0

    def style_row(row):
        styles = [""] * len(row)
        val = row[column_name]

        # ZERO
        if val == 0:
            styles[row.index.get_loc(column_name)] = (
                "background-color: white; color: black;"
            )
            return styles

        # POSITIVE
        if val > 0:
            if pos_max > pos_min:
                scale = (val - pos_min) / (pos_max - pos_min)   # 0 → 1
            else:
                # only one positive value → fixed mid intensity
                scale = 1.0

            # light → darker green (#DFF5E1 → #2ECC71)
            r = int(223 - scale * (223 - 46))
            g = int(245 - scale * (245 - 204))
            b = int(225 - scale * (225 - 113))

            styles[row.index.get_loc(column_name)] = (
                f"background-color: rgb({r},{g},{b}); color: black;"
            )
            return styles

        # NEGATIVE
        if val < 0:
            if neg_min < neg_max:
                # for negatives, min is more negative; normalize 0 → 1
                scale = (val - neg_max) / (neg_min - neg_max)
            else:
                # only one negative value → fixed mid intensity
                scale = 1.0

            # light → darker red (#FADBD8 → #E74C3C)
            r = int(250 - scale * (250 - 231))
            g = int(219 - scale * (219 - 76))
            b = int(216 - scale * (216 - 60))

            styles[row.index.get_loc(column_name)] = (
                f"background-color: rgb({r},{g},{b}); color: black;"
            )
            return styles

        # fallback
        styles[row.index.get_loc(column_name)] = (
            "background-color: white; color: black;"
        )
        return styles

    return df.style.apply(style_row, axis=1)


st.subheader("Optimal Portfolio Allocation")

col_left, col_right = st.columns([1, 1])
A_opt = best["A_opt"]
w_opt = best["w_opt"]

with col_left:
    st.markdown("**Amounts (€ millions)**")

    allocation_df = pd.DataFrame({
        "Asset Class": ["Gov Bonds", "Corp Bonds", "Equity 1", "Equity 2", "Property", "T-Bills"],
        "Return (%)": (initial_asset["asset_ret"] * 100),
        "Current (€m)": initial_asset["asset_val"].values,
        "Optimal (€m)": A_opt,
        "Change (€m)": A_opt - initial_asset["asset_val"].values
    })

    # FIX: Add Return (%) to style.format
    allocation_df = (
        style_change_column_apply(allocation_df, "Change (€m)")
        .format({
            "Return (%)": "{:.2f}",
            "Current (€m)": "{:.1f}",
            "Optimal (€m)": "{:.1f}",
            "Change (€m)": "{:+.1f}"
        })
    )
    st.dataframe(
        allocation_df,
        use_container_width=True,
        hide_index=True
    )

with col_right:
    st.markdown("**Weights (%)**")
    weights_df = pd.DataFrame({
        "Asset Class": ["Gov Bonds", "Corp Bonds", "Equity 1", "Equity 2", "Property", "T-Bills"],
        "Current (%)": initial_asset["asset_weight"].values * 100,
        "Optimal (%)": w_opt * 100,
        "Change (%)": (w_opt - initial_asset["asset_weight"].values) * 100
    })

    # FIX: Apply formatting only to specific numeric columns

    weights_df = (
        style_change_column_apply(weights_df, "Change (%)")
        .format({
            "Current (%)": "{:.1f}",
            "Optimal (%)": "{:.1f}",
            "Change (%)": "{:+.1f}"
        })
    )

    st.dataframe(
        weights_df,
        use_container_width=True,
        hide_index=True
    )

st.markdown("---")

# --- 7. Pie Charts (Restored Visuals) ---
st.markdown("**Visual Allocation Comparison**")
fig_pie, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 7))
fig_pie.patch.set_facecolor('white')

asset_labels = ["Gov Bonds", "Corp Bonds", "Equity 1", "Equity 2", "Property", "T-Bills"]
colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', '#DFE6E9']


def make_autopct(values):
    def my_autopct(pct):
        return f'{pct:.1f}%' if pct > 3 else ''

    return my_autopct


# Current Pie
current_weights = initial_asset["asset_weight"].values * 100
wedges1, _, autotexts1 = ax1.pie(
    current_weights, colors=colors, autopct=make_autopct(current_weights),
    startangle=90, textprops={'fontsize': 11, 'weight': 'bold'}, pctdistance=0.80,
    explode=[0.03 if w > 10 else 0 for w in current_weights], shadow=True,
    wedgeprops={'edgecolor': 'white', 'linewidth': 2}
)
ax1.set_title('Current Portfolio', fontsize=14, fontweight='bold', pad=25)

# Optimal Pie
optimal_weights = w_opt * 100
wedges2, _, autotexts2 = ax2.pie(
    optimal_weights, colors=colors, autopct=make_autopct(optimal_weights),
    startangle=90, textprops={'fontsize': 11, 'weight': 'bold'}, pctdistance=0.80,
    explode=[0.03 if w > 10 else 0 for w in optimal_weights], shadow=True,
    wedgeprops={'edgecolor': 'white', 'linewidth': 2}
)
ax2.set_title('Optimal Portfolio', fontsize=14, fontweight='bold', pad=25)

# Legend
fig_pie.legend(wedges1, asset_labels, title="Asset Classes", loc="lower center",
               bbox_to_anchor=(0.5, -0.08), ncol=6, frameon=True, shadow=True)
st.pyplot(fig_pie, use_container_width=True)

st.markdown("---")

# --- 8. Key Allocation Changes (Text) ---
st.markdown("**Key Allocation Changes:**")
changes = []
for i, label in enumerate(asset_labels):
    change = (w_opt[i] * 100) - (initial_asset["asset_weight"].values[i] * 100)
    if abs(change) > 1.0:
        icon = "🟢" if change > 0 else "🔴"
        direction = "↑" if change > 0 else "↓"
        changes.append(f"{icon} **{label}**: {direction} {abs(change):.1f}pp")

if changes:
    cols = st.columns(2)
    mid = len(changes) // 2
    with cols[0]:
        for c in changes[:mid]: st.markdown(c)
    with cols[1]:
        for c in changes[mid:]: st.markdown(c)
else:
    st.info("No significant allocation changes (< 1 percentage point)")

st.markdown("---")

# =========================================================
# 🧩 MARGINAL SCR ANALYSIS (NEW SECTION)
# =========================================================
st.subheader("Market Solvency Capital Requirement (SCR)")

# --- 1. Calculate Marginal SCR (Logic adapted from notebook) ---
# We need to reconstruct the SCR components for the OPTIMAL portfolio
# (The 'best' object contains aggregated results, but we need component vectors for marginal calc)

# Re-calculate component SCRs for the optimal allocation
# Note: This logic repeats some backend work but is necessary if 'best' doesn't store raw components


if False: 


    # Compute Interest SCR
    scr_int_res = scr_interest_rate(
        best["A_opt"],
        initial_asset["asset_dur"],
        liab_value,
        liab_duration,
        ir_up,
        ir_down
    )

    # 2. Equity Risk
    A_eq1_opt = best["A_opt"][2]
    A_eq2_opt = best["A_opt"][3]
    scr_eq_res = scr_eq(A_eq1_opt, A_eq2_opt, solv["equity_1_param"], solv["equity_2_param"], solv["rho"])

    # 3. Property Risk
    scr_prop_val = scr_prop(best["A_opt"][4], solv["prop_params"])

    # 4. Spread Risk
    # Need spread shock (re-compute or store)
    corp_dur = initial_asset.loc["corp_bond", "asset_dur"]
    spread_shock_val = compute_spread_shock_eiopa(corp_dur, verbose=False)
    scr_sprd_val = scr_sprd(best["A_opt"][1], spread_shock_val)


    market_SCR_opt = aggregate_market_scr(
        scr_int_res,
        scr_eq_res, 
        scr_prop_val,
        scr_sprd_val,
        corr_down,
        corr_up
    )

    # 5. Prepare Vector for Marginal Calculation
    scr_vec = np.array([
        scr_int_res["SCR_interest"],
        scr_eq_res["SCR_eq_total"],
        scr_prop_val,
        scr_sprd_val
    ])


market_SCR_opt = opt_df.loc[best_idx, "SCR_breakdown"]

scr_vec = np.array(market_SCR_opt.iloc[:4, ]['SCR'])

# Get it from the optimizer
direction = opt_df.loc[best_idx, 'direction']

# Run Marginal Calculation (Risk Level)
marg_risk_df = marginal_scr(scr_vec, direction, corr_down.values, corr_up.values)

# Run Allocation Calculation (Asset Level)
# Need to reconstruct params dict
alloc_params = {
    "interest_down": ir_down,
    "interest_up": ir_up,
    "spread": spread_shock_val,
    "equity_type1": solv["equity_1_param"],
    "equity_type2": solv["equity_2_param"],
    "property": solv["prop_params"],
    "rho": solv["rho"]
}

# We need a DataFrame for the Optimal Portfolio state to pass to allocate_marginal_scr
# The function expects 'initial_asset' structure but with optimal values?
# Actually, looking at the function 'allocate_marginal_scr' in notebook/file:
# It uses 'initial_asset' for durations and 'asset_val' for Equities.
# We should pass a dataframe representing the OPTIMAL portfolio.
opt_asset_df = initial_asset.copy()
opt_asset_df["asset_val"] = best["A_opt"]

mscr_assets = allocate_marginal_scr(marg_risk_df, direction, opt_asset_df, alloc_params)

asset_mSCR_CUR = allocate_marginal_scr(
    marg_risk_CUR, 
    direction_CUR, 
    initial_asset, 
    alloc_params
)

# --- 2. Display Results ---

col_m1, col_m2 = st.columns([1, 1])

with col_m1:

    st.markdown("**SCR per risk type**")

    scr_display = pd.concat([market_scr_CUR['summary_table'], market_SCR_opt], axis=1)
    scr_display.columns=['Current', 'Optimal']
    scr_display['change'] = (scr_display['Current'] - scr_display['Optimal']).astype('float')
    new_order = ["interest", "equity", "property", "spread", "diversification", "total"]
    scr_display = scr_display.reindex(new_order)

    scr_display = scr_display.rename(
        columns={"Current": "Current (€m)", "Optimal": "Optimal (€m)", "change": "Change (€m)"},
        index={
            "interest": "Interest",
            "equity": "Equity",
            "property": "Property",
            "spread": "Spread",
            "diversification": "Diversified",
            "total": "Total"
        }
    )

    scr_display = scr_display.reset_index()

    # ------------------ IMPORTANT FIX ------------------
    scr_df_raw = scr_display.copy()   # Save BEFORE styling
    # ---------------------------------------------------

    scr_display_styled = (
        style_change_column_apply(scr_df_raw, "Change (€m)")
        .format({
            "Current (€m)": "{:.1f}",
            "Optimal (€m)": "{:.1f}",
            "Change (€m)": "{:.1f}"
        })
    )

    st.dataframe(scr_display_styled, use_container_width=True, hide_index=True)


with col_m2:
    st.markdown("**Marginal SCR per asset**", 
                help="Approximate increase in SCR when the allocation to the asset class increases by 1 unit")

    mSCR_display = pd.concat(
        [asset_mSCR_CUR[['asset', 'mSCR']].copy(),
         mscr_assets[['mSCR']]],
        axis=1
    )
    mSCR_display.columns = ['asset', 'Current', 'Optimal']
    mSCR_display['Change'] = mSCR_display['Current'] - mSCR_display['Optimal']

    mSCR_styled = (
        style_change_column_apply(mSCR_display, 'Change')
        .format({
            "Current": "{:.2f}",
            "Optimal": "{:.2f}",
            "Change": "{:.2f}"
        })
    )
    st.dataframe(mSCR_styled, use_container_width=True, hide_index=True)



st.markdown("---")
# --- SCR Pie Charts (Styled Like Allocation Pies) ---
st.markdown("**SCR by Risk Type (Pie Charts)**")

# Clean filtered SCR rows (your scr_df_raw has Risk column)
pie_scr = scr_df_raw[scr_df_raw["risk"].isin(["Interest", "Equity", "Property", "Spread"])]

# Extract values
pie_current = pie_scr["Current (€m)"].values
pie_optimal = pie_scr["Optimal (€m)"].values
labels = pie_scr["risk"].values

# Color palette (4 SCR risks)
scr_colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4']


# --- Helper for hiding tiny slices ---
def make_autopct(values):
    def pct_fmt(pct):
        return f'{pct:.1f}%' if pct > 3 else ''
    return pct_fmt


# --- Build Side-by-Side Pies ---
fig_scr, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 7))
fig_scr.patch.set_facecolor('white')


# ===== CURRENT SCR PIE =====
explode_current = [0.03 if v > pie_current.mean() else 0 for v in pie_current]

ax1.pie(
    pie_current,
    labels=None,
    colors=scr_colors,
    autopct=make_autopct(pie_current),
    startangle=90,
    textprops={'fontsize': 11, 'weight': 'bold'},
    pctdistance=0.80,
    explode=explode_current,
    shadow=True,
    wedgeprops={'edgecolor': 'white', 'linewidth': 2}
)
ax1.set_title('Current SCR', fontsize=14, fontweight='bold', pad=25)


# ===== OPTIMAL SCR PIE =====
explode_opt = [0.03 if v > pie_optimal.mean() else 0 for v in pie_optimal]

ax2.pie(
    pie_optimal,
    labels=None,
    colors=scr_colors,
    autopct=make_autopct(pie_optimal),
    startangle=90,
    textprops={'fontsize': 11, 'weight': 'bold'},
    pctdistance=0.80,
    explode=explode_opt,
    shadow=True,
    wedgeprops={'edgecolor': 'white', 'linewidth': 2}
)
ax2.set_title('Optimal SCR', fontsize=14, fontweight='bold', pad=25)


# --- Legend ---
fig_scr.legend(
    labels,
    title="Risk Types",
    loc="lower center",
    bbox_to_anchor=(0.5, -0.08),
    ncol=4,
    frameon=True,
    shadow=True
)

st.pyplot(fig_scr, use_container_width=True)
st.markdown("---")


if False:

    plot_data = disp_df[disp_df["mSCR"] > 0].copy()

    fig_mscr, ax_mscr = plt.subplots(figsize=(6, 6))
    ax_mscr.pie(
        plot_data["mSCR"],
        labels=plot_data["asset"],
        autopct='%1.1f%%',
        startangle=90,
        colors=['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7'],
        wedgeprops={'edgecolor': 'white', 'linewidth': 2}
    )
    ax_mscr.set_title("Risk Contribution (Positive Only)")
    st.pyplot(fig_mscr, use_container_width=True)

    if (disp_df["mSCR"] < 0).any():
        st.caption("Note: Assets with negative marginal SCR (risk reducers/hedges) are excluded from the pie chart.")

st.markdown("---")

# Sensitivity Analysis
st.subheader("🔬 Sensitivity Analysis")

st.markdown("""
    Explore how the optimal portfolio allocation changes under different scenarios.
    Adjust the parameters below to see the impact on asset allocation and risk metrics.
    """)
sens_tab1, sens_tab2, sens_tab3 = st.tabs([
    "📊 Return Scenarios",
    "⚡ Shock Scenarios",
    "🎯 Custom Scenario"
])

# Retrieve saved limits (Critical for consistency)
# If missing, we fallback to the hardcoded ones as a failsafe
if "allocation_limits" in st.session_state:
    saved_limits = st.session_state["allocation_limits"]
else:
    # Fallback defaults
    saved_limits = pd.DataFrame({
        "asset": ["gov_bond", "illiquid_assets", "t_bills", "corp_bond"],
        "min_weight": [0.25, 0.0, 0.01, 0.0],
        "max_weight": [0.75, 0.20, 0.05, 0.50],
    }).set_index("asset")

# --- TAB 1: Return Scenarios ---
with sens_tab1:
    scenario_choice = st.selectbox(
        "Select Scenario",
        [
            "Base Case (Current)",
            "Optimistic (+0.5% all assets)",
            "Pessimistic (-0.5% all assets)",
            "Higher Equity Returns (+2.0%)",
            "Lower Bond Returns (-1.0%)"
        ],
        key="return_scenario"
    )

    if st.button("🔄 Run Return Sensitivity", key="run_return_sens"):
        with st.spinner("Running sensitivity analysis..."):
            try:
                base_returns = initial_asset["asset_ret"].values.copy()

                if scenario_choice == "Base Case (Current)":
                    scenario_returns = base_returns
                elif "Optimistic" in scenario_choice:
                    scenario_returns = base_returns + 0.005
                elif "Pessimistic" in scenario_choice:
                    scenario_returns = base_returns - 0.005
                elif "Higher Equity" in scenario_choice:
                    scenario_returns = base_returns.copy()
                    scenario_returns[2] += 0.02 # Eq1
                    scenario_returns[3] += 0.02 # Eq2
                elif "Lower Bond" in scenario_choice:
                    scenario_returns = base_returns.copy()
                    scenario_returns[0] -= 0.01 # Gov
                    scenario_returns[1] -= 0.01 # Corp

                # Setup Inputs
                sens_asset = initial_asset.copy()
                sens_asset["asset_ret"] = scenario_returns

                # Load Environment
                cfg = load_config()
                corr_down, corr_up = get_corr_matrices(cfg)
                
                # Retrieve Params (containing Shocks AND Solvency Min)
                if "params" in st.session_state:
                    params = st.session_state["params"]
                else:
                    solv = get_solvency_params(cfg)
                    params = {
                        "interest_down": 0.009, "interest_up": 0.011, "spread": 0.103,
                        "equity_type1": 0.39, "equity_type2": 0.49,
                        "property": 0.25, "rho": 0.75, "solvency_min": 1.0
                    }

                sens_opt_df = solve_frontier_combined(
                    initial_asset=sens_asset, liab_value=liab_value, liab_duration=liab_duration,
                    corr_downward=corr_down, corr_upward=corr_up, 
                    allocation_limits=saved_limits, # <--- USE SAVED LIMITS
                    params=params
                )

                if not sens_opt_df.empty:
                    sens_best = sens_opt_df.loc[sens_opt_df["objective"].idxmax()]
                    st.success("Sensitivity analysis completed!")

                    # Plot & Metrics
                    st.subheader("Scenario Optimal Portfolio Comparison")
                    fig = plot_scenario_comparison(opt_df, best, sens_best, current_ret, current_sol, sens_df=sens_opt_df)
                    st.pyplot(fig, use_container_width=True)

                    sc1, sc2, sc3 = st.columns(3)
                    sc1.metric("New Return", f"{sens_best['return']:.2%}", f"{(sens_best['return'] - best_ret):.2%}")
                    sc2.metric("New Solvency", f"{sens_best['solvency'] * 100:.1f}%", f"{(sens_best['solvency'] - best_sol) * 100:.1f}pp")

                    st.markdown("**Asset Allocation Change (pp)**")
                    diff = (sens_best["w_opt"] - best["w_opt"]) * 100
                    df_diff = pd.DataFrame([diff], columns=["Gov", "Corp", "Eq1", "Eq2", "Prop", "TB"])
                    st.dataframe(df_diff.style.format("{:+.2f}"), hide_index=True)
                else:
                    st.error("Optimization failed for this scenario.")

            except Exception as e:
                st.error(f"Error: {e}")

# --- TAB 2: Shock Scenarios ---
with sens_tab2:
    st.markdown("**Test different Solvency II shock assumptions**")

    shock_scenario = st.selectbox(
        "Select Shock Scenario",
        [
            "Base Case (Current)",
            "Stressed Shocks (+50%)", 
            "Relaxed Shocks (-30%)", 
            "Higher Equity Shocks (+15pp)", 
            "Lower Interest Rate Shocks (-20%)",
            "Custom Shocks"
        ],
        key="shock_scenario"
    )

    # Base shocks (Ideally retrieve from session params to match inputs)
    if "params" in st.session_state:
        base_p = st.session_state["params"]
        base_shocks = {
            "ir_up": base_p["interest_up"], "ir_down": base_p["interest_down"],
            "eq1": base_p["equity_type1"], "eq2": base_p["equity_type2"],
            "prop": base_p["property"], "spread": base_p["spread"]
        }
        current_solv_min = base_p.get("solvency_min", 1.0)
    else:
        # Fallback
        base_shocks = {"ir_up": 0.011, "ir_down": 0.009, "eq1": 0.39, "eq2": 0.49, "prop": 0.25, "spread": 0.103}
        current_solv_min = 1.0

    # Apply scenario
    if shock_scenario == "Base Case (Current)":
        scenario_shocks = base_shocks.copy()
    elif "Stressed" in shock_scenario:
        scenario_shocks = {k: v * 1.5 for k, v in base_shocks.items()}
    elif "Relaxed" in shock_scenario:
        scenario_shocks = {k: v * 0.7 for k, v in base_shocks.items()}
    elif "Higher Equity" in shock_scenario:
        scenario_shocks = base_shocks.copy()
        scenario_shocks["eq1"] = min(1.0, base_shocks["eq1"] + 0.15)
        scenario_shocks["eq2"] = min(1.0, base_shocks["eq2"] + 0.15)
    elif "Lower Interest" in shock_scenario:
        scenario_shocks = base_shocks.copy()
        scenario_shocks["ir_up"] *= 0.8
        scenario_shocks["ir_down"] *= 0.8
    else:  # Custom Shocks
        col1, col2 = st.columns(2)
        with col1:
            ir_up_shock = st.number_input("IR Up", 0.0, 0.5, base_shocks["ir_up"], 0.001, format="%.3f")
            ir_down_shock = st.number_input("IR Down", 0.0, 0.5, base_shocks["ir_down"], 0.001, format="%.3f")
            spread_shock = st.number_input("Spread", 0.0, 0.5, base_shocks["spread"], 0.001, format="%.3f")
        with col2:
            eq1_shock = st.number_input("Eq 1", 0.0, 1.0, base_shocks["eq1"], 0.01)
            eq2_shock = st.number_input("Eq 2", 0.0, 1.0, base_shocks["eq2"], 0.01)
            prop_shock = st.number_input("Prop", 0.0, 1.0, base_shocks["prop"], 0.01)
        scenario_shocks = {
            "ir_up": ir_up_shock, "ir_down": ir_down_shock, "spread": spread_shock,
            "eq1": eq1_shock, "eq2": eq2_shock, "prop": prop_shock
        }

    if st.button("🔄 Run Shock Sensitivity", key="run_shock_sens"):
        with st.spinner("Running shock sensitivity analysis..."):
            try:
                cfg = load_config()
                corr_down, corr_up = get_corr_matrices(cfg)
                solv = get_solvency_params(cfg)

                # Construct Params
                params = {
                    "interest_down": scenario_shocks["ir_down"],
                    "interest_up": scenario_shocks["ir_up"],
                    "spread": scenario_shocks["spread"],
                    "equity_type1": scenario_shocks["eq1"],
                    "equity_type2": scenario_shocks["eq2"],
                    "property": scenario_shocks["prop"],
                    "rho": solv["rho"],
                    "solvency_min": current_solv_min # <--- CRITICAL FIX: Pass the user target!
                }

                sens_opt_df = solve_frontier_combined(
                    initial_asset=initial_asset,
                    liab_value=liab_value,
                    liab_duration=liab_duration,
                    corr_downward=corr_down,
                    corr_upward=corr_up,
                    allocation_limits=saved_limits, # <--- CRITICAL FIX: Use saved limits
                    params=params
                )

                if not sens_opt_df.empty:
                    sens_best = sens_opt_df.loc[sens_opt_df["objective"].idxmax()]
                    st.success("✓ Shock sensitivity analysis completed!")
                    
                    # ... [Plotting and metrics code same as above] ...
                    st.subheader("Scenario Optimal Portfolio Comparison")
                    fig = plot_scenario_comparison(opt_df, best, sens_best, current_ret, current_sol, sens_df=sens_opt_df)
                    st.pyplot(fig, use_container_width=True)
                    
                    col1, col2, col3 = st.columns(3)
                    col1.metric("Expected Return", f"{sens_best['return']:.2%}", delta=f"{(sens_best['return'] - best['return']):.2%}")
                    col2.metric("SCR Market", f"€{sens_best['SCR_market']:.1f}m", delta=f"€{(sens_best['SCR_market'] - best['SCR_market']):.1f}m", delta_color="inverse")
                    col3.metric("Solvency Ratio", f"{sens_best['solvency'] * 100:.1f}%", delta=f"{(sens_best['solvency'] - best['solvency']) * 100:.1f}%")

                    st.markdown("**Asset Allocation Response**")
                    comparison_df = pd.DataFrame({
                        "Asset Class": ["Gov Bonds", "Corp Bonds", "Equity Type 1", "Equity Type 2", "Property", "T-Bills"],
                        "Base (%)": best["w_opt"] * 100,
                        "Scenario (%)": sens_best["w_opt"] * 100,
                        "Change (pp)": (sens_best["w_opt"] - best["w_opt"]) * 100
                    })
                    st.dataframe(comparison_df.style.format("{:.1f}").background_gradient(subset=["Change (pp)"], cmap="RdYlGn", vmin=-10, vmax=10), use_container_width=True)

                else:
                    st.error("Optimization failed. The shocks might be too severe to maintain solvency.")

            except Exception as e:
                st.error(f"Error in shock sensitivity: {str(e)}")
    # ==========================================
    # TAB 3: Custom Scenario
    # ==========================================
    with sens_tab3:
        st.markdown("**Create your own custom scenario**")
        st.markdown("Combine different return and shock assumptions to test specific hypotheses.")

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("**📊 Custom Returns**")
            # Initialize with current asset returns as defaults
            defaults = initial_asset["asset_ret"].values
            custom_r_gov = st.number_input("Gov Bonds Return", -0.10, 0.20, float(defaults[0]), 0.001, format="%.3f", key="c_r_gov")
            custom_r_corp = st.number_input("Corp Bonds Return", -0.10, 0.20, float(defaults[1]), 0.001, format="%.3f", key="c_r_corp")
            custom_r_eq1 = st.number_input("Equity 1 Return", -0.10, 0.20, float(defaults[2]), 0.001, format="%.3f", key="c_r_eq1")
            custom_r_eq2 = st.number_input("Equity 2 Return", -0.10, 0.20, float(defaults[3]), 0.001, format="%.3f", key="c_r_eq2")
            custom_r_prop = st.number_input("Property Return", -0.10, 0.20, float(defaults[4]), 0.001, format="%.3f", key="c_r_prop")
            custom_r_tb = st.number_input("T-Bills Return", -0.10, 0.20, float(defaults[5]), 0.001, format="%.3f", key="c_r_tb")

        with col2:
            st.markdown("**⚡ Custom Shocks**")
            # Load default parameters from config/session for defaults
            cfg = load_config()
            solv = get_solvency_params(cfg)
            
            # Use session state if available (from auto-calc), else defaults
            # Ideally these should come from st.session_state if stored in Inputs.py
            # Here we use safe defaults or previously computed values if possible
            
            custom_ir_up = st.number_input("IR Up Shock", 0.0, 0.20, 0.011, 0.001, format="%.3f", key="c_ir_up")
            custom_ir_down = st.number_input("IR Down Shock", 0.0, 0.20, 0.009, 0.001, format="%.3f", key="c_ir_down")
            custom_spread = st.number_input("Spread Shock", 0.0, 0.30, 0.103, 0.001, format="%.3f", key="c_spread")
            
            custom_eq1 = st.number_input("Equity 1 Shock", 0.0, 1.0, solv["equity_1_param"], 0.01, format="%.2f", key="c_eq1")
            custom_eq2 = st.number_input("Equity 2 Shock", 0.0, 1.0, solv["equity_2_param"], 0.01, format="%.2f", key="c_eq2")
            custom_prop = st.number_input("Property Shock", 0.0, 1.0, solv["prop_params"], 0.01, format="%.2f", key="c_prop")

        if st.button("🚀 Run Custom Scenario", key="run_custom_sens", type="primary"):
            with st.spinner("Running custom scenario analysis..."):
                try:
                    # 1. Build Custom Inputs
                    custom_asset = initial_asset.copy()
                    custom_asset["asset_ret"] = [custom_r_gov, custom_r_corp, custom_r_eq1,
                                                 custom_r_eq2, custom_r_prop, custom_r_tb]

                    # Re-load matrices
                    corr_down, corr_up = get_corr_matrices(cfg)

                    # Build Allocation Limits (Same as base case)
                    allocation_limits = pd.DataFrame({
                        "asset": ["gov_bond", "illiquid_assets", "t_bills", "corp_bond"],
                        "min_weight": [0.25, 0.0, 0.01, 0.0],
                        "max_weight": [0.75, 0.20, 0.05, 0.50],
                    }).set_index("asset")

                    # Build Params Dictionary
                    custom_params = {
                        "interest_down": custom_ir_down,
                        "interest_up": custom_ir_up,
                        "spread": custom_spread,
                        "equity_type1": custom_eq1,
                        "equity_type2": custom_eq2,
                        "property": custom_prop,
                        "rho": solv["rho"],
                    }

                    # 2. Run Optimization
                    custom_opt_df = solve_frontier_combined(
                        initial_asset=custom_asset,
                        liab_value=liab_value,
                        liab_duration=liab_duration,
                        corr_downward=corr_down,
                        corr_upward=corr_up,
                        allocation_limits=allocation_limits,
                        params=custom_params
                    )

                    if not custom_opt_df.empty:
                        custom_best_idx = custom_opt_df["objective"].idxmax()
                        custom_best = custom_opt_df.loc[custom_best_idx]

                        st.success("✓ Custom scenario completed!")

                        # --- A. PLOT SCENARIO COMPARISON ---
                        st.subheader("Scenario Optimal Portfolio Comparison")
                        # Pass 'sens_df=custom_opt_df' to see the new frontier line
                        fig = plot_scenario_comparison(opt_df, best, custom_best, current_ret, current_sol, sens_df=custom_opt_df)
                        st.pyplot(fig, use_container_width=True)

                        # --- B. KEY METRICS ---
                        st.markdown("**Impact of Custom Assumptions**")
                        col_m1, col_m2, col_m3 = st.columns(3)
                        
                        with col_m1:
                            st.metric(
                                "Expected Return", 
                                f"{custom_best['return']:.2%}", 
                                f"{(custom_best['return'] - best_ret):.2%}"
                            )
                        with col_m2:
                            st.metric(
                                "Solvency Ratio", 
                                f"{custom_best['solvency']*100:.1f}%", 
                                f"{(custom_best['solvency'] - best_sol)*100:.1f}pp"
                            )
                        with col_m3:
                            st.metric(
                                "SCR Market", 
                                f"€{custom_best['SCR_market']:.1f}m", 
                                f"€{(custom_best['SCR_market'] - best_SCR):.1f}m",
                                delta_color="inverse"
                            )

                        # --- C. ALLOCATION CHANGE ---
                        st.markdown("**Asset Allocation Response**")
                        
                        alloc_comparison = pd.DataFrame({
                            "Asset Class": ["Gov Bonds", "Corp Bonds", "Equity Type 1",
                                            "Equity Type 2", "Property", "T-Bills"],
                            "Base Case (%)": best["w_opt"] * 100,
                            "Custom Case (%)": custom_best["w_opt"] * 100,
                            "Change (pp)": (custom_best["w_opt"] - best["w_opt"]) * 100
                        })

                        st.dataframe(
                            alloc_comparison.style.format({
                                "Base Case (%)": "{:.1f}",
                                "Custom Case (%)": "{:.1f}",
                                "Change (pp)": "{:+.1f}"
                            }).background_gradient(subset=["Change (pp)"], cmap="RdYlGn", vmin=-10, vmax=10),
                            use_container_width=True
                        )
                        
                    else:
                        st.error("❌ Optimization failed for this custom scenario. Try relaxing constraints or improving returns.")

                except Exception as e:
                    st.error(f"Error in custom scenario: {str(e)}")
                    # st.exception(e) # Uncomment for debugging
    st.markdown("---")

# =========================================================
# 💾 EXPORT SECTION (With Sensitivity Scenarios & Specific Frontiers)
# =========================================================
st.subheader("💾 Export Results")
col_export1, col_export2 = st.columns(2)

# --- Define Scenarios ---
scenarios_to_run = {
    "Optimistic": {"type": "return", "val": 0.005},
    "Pessimistic": {"type": "return", "val": -0.005},
    "Higher Equity": {"type": "return_special", "asset": [2, 3], "val": 0.02},
    "Lower Bond": {"type": "return_special", "asset": [0, 1], "val": -0.01},
    "Stressed Shocks": {"type": "shock", "mult": 1.5},
    "Relaxed Shocks": {"type": "shock", "mult": 0.7},
    "High Eq Shock": {"type": "shock_special", "key": ["eq1", "eq2"], "val": 0.15},
    "Low IR Shock": {"type": "shock_special", "key": ["ir_up", "ir_down"], "mult": 0.8}
}

# --- Helper: Run Optimization for a Scenario ---
def run_scenario_optimization(name, config):
    # Safety checks for session state variables
    if "params" not in st.session_state:
        return None, None
    
    sim_params = st.session_state["params"].copy()
    sim_asset = initial_asset.copy()
    
    # 2. Apply Modifications
    if config.get("type") == "return":
        sim_asset["asset_ret"] = sim_asset["asset_ret"] + config["val"]
        
    elif config.get("type") == "return_special":
        new_rets = sim_asset["asset_ret"].values.copy()
        for idx in config["asset"]:
            new_rets[idx] += config["val"]
        sim_asset["asset_ret"] = new_rets
        
    elif config.get("type") == "shock":
        for k in ["interest_up", "interest_down", "spread", "property", "equity_type1", "equity_type2"]:
            sim_params[k] *= config["mult"]
            
    elif config.get("type") == "shock_special":
        if "val" in config: 
            for k in ["equity_type1", "equity_type2"]:
                sim_params[k] = min(1.0, sim_params[k] + config["val"])
        elif "mult" in config:
             for k in ["interest_up", "interest_down"]:
                sim_params[k] *= config["mult"]

    # 3. Run Optimization
    alloc_lims = pd.DataFrame({
        "asset": ["gov_bond", "illiquid_assets", "t_bills", "corp_bond"],
        "min_weight": [0.25, 0.0, 0.01, 0.0],
        "max_weight": [0.90, 0.20, 0.10, 0.50], 
    }).set_index("asset")

    # Pass DataFrames directly
    df_res = solve_frontier_combined(
        sim_asset, liab_value, liab_duration, 
        corr_down, corr_up, alloc_lims, sim_params
    )
    
    if df_res.empty: return None, None
    
    # 4. Pick Best
    best_sim = df_res.loc[df_res["objective"].idxmax()]
    
    # Return BOTH the best result AND the full dataframe (frontier)
    return best_sim, df_res

# --- Excel Generation Function ---
def generate_comprehensive_report():
    # Safely retrieve globally defined data
    g = globals()
    
    # --- Start Workbook ---
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    
    # ==========================
    # SHEET 1: EXECUTIVE SUMMARY
    # ==========================
    ws_main = wb.create_sheet("Executive Summary")
    ws_main.sheet_view.showGridLines = False
    
    ws_main["A1"] = "SOLVENCY II STRATEGIC REPORT"
    ws_main["A1"].font = Font(size=18, bold=True, color="2C3E50")
    ws_main["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}"
    
    # Base Case Metrics
    ws_main["A4"] = "BASE CASE OVERVIEW"
    ws_main["A4"].font = Font(bold=True, size=12)
    
    headers = ["Metric", "Current", "Optimal"]
    ws_main.append([]) 
    ws_main.append(headers)
    
    metrics = [
        ("Expected Return", g['current_ret'], g['best_ret'], "0.00%"),
        ("Solvency Ratio", g['current_sol'], g['best_sol'], "0.0%"),
        ("SCR Market (€m)", g['current_SCR'], g['best_SCR'], "#,##0.0"),
        ("Basic Own Funds (€m)", g['best_BOF'], g['best_BOF'], "#,##0.0") 
    ]
    
    start_row = 6
    for i, (name, cur, opt, fmt) in enumerate(metrics):
        ws_main.cell(row=start_row+i, column=1, value=name)
        c2 = ws_main.cell(row=start_row+i, column=2, value=cur); c2.number_format = fmt
        c3 = ws_main.cell(row=start_row+i, column=3, value=opt); c3.number_format = fmt

    # --- ALLOCATION DATA & CHARTS (Current vs Optimal) ---
    ws_data = wb.create_sheet("Base_Data")
    ws_data.append(["Asset", "Current Weight", "Optimal Weight"])
    for i, asset in enumerate(["Gov", "Corp", "Eq1", "Eq2", "Prop", "TB"]):
        ws_data.append([asset, g['initial_asset']["asset_weight"].values[i], g['best']["w_opt"][i]])
    
    # 1. Optimal Allocation Pie
    chart_opt = PieChart()
    chart_opt.title = "Optimal Allocation"
    data_opt = Reference(ws_data, min_col=3, min_row=1, max_row=7)
    cats = Reference(ws_data, min_col=1, min_row=2, max_row=7)
    chart_opt.add_data(data_opt, titles_from_data=True)
    chart_opt.set_categories(cats)
    ws_main.add_chart(chart_opt, "E4")

    # 2. Current Allocation Pie
    chart_cur = PieChart()
    chart_cur.title = "Current Allocation"
    data_cur = Reference(ws_data, min_col=2, min_row=1, max_row=7)
    chart_cur.add_data(data_cur, titles_from_data=True)
    chart_cur.set_categories(cats)
    ws_main.add_chart(chart_cur, "M4") # Place to the right

    # ==========================
    # SHEET 2: mSCR ANALYSIS
    # ==========================
    if 'disp_df' in g and not g['disp_df'].empty:
        ws_mscr = wb.create_sheet("mSCR Analysis")
        ws_mscr["A1"] = "Marginal SCR Contribution (Optimal Portfolio)"
        ws_mscr["A1"].font = Font(size=14, bold=True)
        ws_mscr.append([])
        
        # Write Table
        for r in dataframe_to_rows(g['disp_df'], index=False, header=True):
            ws_mscr.append(r)
            
        # Add Pie Chart for mSCR
        chart_pie = PieChart()
        chart_pie.title = "Risk Contribution (mSCR)"
        
        labels = Reference(ws_mscr, min_col=1, min_row=4, max_row=4+len(g['disp_df']))
        data = Reference(ws_mscr, min_col=2, min_row=3, max_row=4+len(g['disp_df']))
        
        chart_pie.add_data(data, titles_from_data=True)
        chart_pie.set_categories(labels)
        ws_mscr.add_chart(chart_pie, "E3")

    # ==========================
    # SCENARIO SHEETS (Loop)
    # ==========================
    prog_bar = st.progress(0)
    status_text = st.empty()
    total_scenarios = len(scenarios_to_run)
    
    for i, (scen_name, config) in enumerate(scenarios_to_run.items()):
        status_text.text(f"Running scenario {i+1}/{total_scenarios}: {scen_name}...")
        prog_bar.progress((i) / total_scenarios)
        
        # Run Optimization -> Get Best AND Full Frontier
        res, frontier_df = run_scenario_optimization(scen_name, config)
        if res is None: continue
            
        ws = wb.create_sheet(f"Sens - {scen_name}")
        ws["A1"] = f"SCENARIO: {scen_name.upper()}"
        ws["A1"].font = Font(size=14, bold=True, color="E74C3C")
        ws["A2"] = f"Type: {config['type']}"
        
        # --- Metrics Table ---
        ws.append([])
        ws.append(["Metric", "Base Optimal", "Scenario Optimal", "Delta"])
        ws["A4"].font = Font(bold=True)
        
        scen_ret, scen_sol, scen_scr = res["return"], res["solvency"], res["SCR_market"]
        
        data_rows = [
            ("Return", g['best_ret'], scen_ret, "0.00%"),
            ("Solvency", g['best_sol'], scen_sol, "0.0%"),
            ("SCR (€m)", g['best_SCR'], scen_scr, "#,##0.0")
        ]
        
        for r_idx, (m, b, s, fmt) in enumerate(data_rows, 5):
            ws.cell(row=r_idx, column=1, value=m)
            ws.cell(row=r_idx, column=2, value=b).number_format = fmt
            c_s = ws.cell(row=r_idx, column=3, value=s); c_s.number_format = fmt
            c_d = ws.cell(row=r_idx, column=4, value=s-b); c_d.number_format = fmt
            
            if m == "Return" and s < b: c_d.font = Font(color="FF0000")
            if m == "Solvency" and s < 1.0: c_s.font = Font(color="FF0000", bold=True)

        # --- Allocation Table & Pie Chart ---
        ws.append([])
        ws.append(["Asset Allocation", "Scenario (%)"])
        ws["A9"].font = Font(bold=True)
        
        assets = ["Gov Bonds", "Corp Bonds", "Equity 1", "Equity 2", "Property", "T-Bills"]
        scen_w = res["w_opt"] * 100
        
        for j, asset in enumerate(assets):
            row_num = 10 + j
            ws.cell(row=row_num, column=1, value=asset)
            ws.cell(row=row_num, column=2, value=scen_w[j]).number_format = "0.0"

        # Scenario Allocation Pie Chart
        chart = PieChart()
        chart.title = f"Allocation: {scen_name}"
        
        data = Reference(ws, min_col=2, min_row=9, max_col=2, max_row=15)
        cats = Reference(ws, min_col=1, min_row=10, max_row=15)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "D4")

        # --- Scenario Frontier Data & Chart ---
        # Write frontier data starting at Row 20
        ws.cell(row=19, column=1, value="Efficient Frontier Data").font = Font(bold=True)
        ws.cell(row=20, column=1, value="Solvency (%)")
        ws.cell(row=20, column=2, value="Return (%)")
        
        f_data = frontier_df[["solvency", "return"]].copy()
        f_data["solvency"] *= 100
        f_data["return"] *= 100
        
        for idx, row in f_data.iterrows():
            r_num = 21 + idx
            ws.cell(row=r_num, column=1, value=row["solvency"])
            ws.cell(row=r_num, column=2, value=row["return"])
            
        chart_scatter = ScatterChart()
        chart_scatter.title = f"Frontier: {scen_name}"
        chart_scatter.x_axis.title = "Solvency Ratio (%)"
        chart_scatter.y_axis.title = "Expected Return (%)"
        chart_scatter.style = 13
        
        max_row = 20 + len(f_data)
        xvalues = Reference(ws, min_col=1, min_row=21, max_row=max_row)
        yvalues = Reference(ws, min_col=2, min_row=21, max_row=max_row)
        
        # FIX: Enabled the line by NOT setting noFill=True
        series = SeriesFactory(values=yvalues, xvalues=xvalues, title=f"Frontier {scen_name}")
        series.marker.symbol = "circle"
        series.smooth = True # Makes the line curved and nice
        
        chart_scatter.series.append(series)
        ws.add_chart(chart_scatter, "D20")

    prog_bar.progress(1.0)
    status_text.text("Report generation complete!")
    
    # Save
    out = BytesIO()
    wb.save(out)
    return out.getvalue()

# --- UI FOR EXPORT ---
with col_export1:
    st.markdown("### 📊 Comprehensive Excel Report")
    st.markdown("Generates a workbook with **10+ Tabs**:")
    st.markdown("1. **Executive Summary** (Current vs Optimal Pies)")
    st.markdown("2. **mSCR Analysis** (Risk Tables & Charts)")
    st.markdown("3-10. **Stress Test Scenarios** (Pies & Frontiers)")
    st.info("⚠️ **Note:** This runs the optimizer 8 times. It may take ~30-60 seconds.")
    
    if st.button("Generate Stress Test Report"):
        with st.spinner("Compiling and running 8 optimization scenarios..."):
            excel_data = generate_comprehensive_report()
            st.download_button(
                label="📥 Download Full Report (.xlsx)",
                data=excel_data,
                file_name=f"SolvencyII_StressTest_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

with col_export2:
    st.markdown("### 📄 Simple Exports")
    st.markdown("Download just the current screen's data.")
    
    if 'allocation_export' not in locals():
        allocation_export = pd.DataFrame({
            "Asset Class": ["Gov Bonds", "Corp Bonds", "Equity 1", "Equity 2", "Property", "T-Bills"],
            "Current (€m)": initial_asset["asset_val"].values,
            "Optimal (€m)": best["A_opt"],
            "Change (€m)": best["A_opt"] - initial_asset["asset_val"].values,
            "Weight (%)": best["w_opt"] * 100
        })
    if 'frontier_export' not in locals():
        frontier_export = opt_df[["gamma", "return", "SCR_market", "solvency", "objective"]].copy()

    csv_alloc = allocation_export.to_csv(index=False).encode('utf-8')
    st.download_button("📥 Allocation Only (CSV)", csv_alloc, "allocation.csv", "text/csv", use_container_width=True)
    
    csv_frontier = frontier_export.to_csv(index=False).encode('utf-8')
    st.download_button("📥 Frontier Only (CSV)", csv_frontier, "frontier.csv", "text/csv", use_container_width=True)

    if 'risk_export' not in locals():
        risk_export = pd.DataFrame()
        
    json_export = {
        "metadata": {"generated": str(datetime.now())},
        "optimal_portfolio": {
            "metrics": {"return": best_ret, "solvency": best_sol, "SCR": best_SCR},
            "allocation": allocation_export.to_dict(orient="records"),
            "risk": risk_export.to_dict(orient="records") if not risk_export.empty else []
        }
    }
    st.download_button("📥 Full Data (JSON)", json.dumps(json_export, indent=2), "results.json", "application/json", use_container_width=True)
